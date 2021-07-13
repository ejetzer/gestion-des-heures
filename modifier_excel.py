#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""modifier_excel.py - Mettre à jour les feuilles de temps directement.

Created on Sun Apr 25 09:03:13 2021

@author: emilejetzer
"""

import openpyxl as xl
# from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.utils.dataframe import dataframe_to_rows

from pathlib import Path
import pandas as pd
import itertools
from datetime import datetime as Dt
from typing import Any, Iterator, Union

from entrer_les_heures import extraire, formater,\
    répartition, compte_des_heures, archiver


racine = Path.home() / 'Documents' / 'Polytechnique' / 'Heures'
CalendrierTemps = 'CalendrierTemps-2021-2022 Émile_Jetzer.xlsx'
TempsTechnicien = 'Feuille de temps 2021-2022 Émile.xlsx'
# TempsAtelier = 'TempsAtelier - Émile.xlsx'


def màj_temps_technicien(nouvelles_données: pd.DataFrame,
                         fichier: Path = racine / TempsTechnicien,
                         nom_feuille: str = 'feuille de temps'):
    noms_de_colonnes = 'A1:J1'
    rangée_min = 2
    colonne_max = 10

    cahier = xl.load_workbook(fichier)
    feuille = cahier[nom_feuille]
    colonnes = [col.value for col in feuille[noms_de_colonnes][0]]
    valeurs: dict[str, list[Any]] = {c: [] for c in colonnes}
    for ligne in feuille.iter_rows(min_row=rangée_min, max_col=colonne_max):
        cellules = [cel.value for cel in ligne]
        if not any(cellules):
            break
        for colonne, cellule in zip(colonnes, cellules):
            valeurs[colonne].append(cellule)
    tableau = pd.DataFrame(valeurs)

    temps_technicien = nouvelles_données\
        .loc[:, ['Technicien',
                 'Payeur',
                 'Date',
                 'Description des travaux effectués',
                 'Demandeur',
                 "Nbr total d'heures",
                 "Nbr d'heures \npassées dans l'atelier",
                 'Précision si pour département',
                 'Taux facturé',
                 'Facturé',
                 'Autres']]

    tableau = tableau.append(temps_technicien)

    for i, ligne in enumerate(dataframe_to_rows(tableau,
                                                index=False,
                                                header=False)):
        for col, cel in zip('ABCDEFGH', ligne):
            feuille[f'{col}{rangée_min+i}'] = cel

    cahier.save(fichier)

    return tableau


# def màj_temps_atelier(temps_atelier: pd.DataFrame,
#                       fichier: Path = racine / TempsAtelier):
#     cahier = xl.load_workbook(fichier)
#     tableaux = temps_atelier.groupby('Payeur')

#     for payeur, heures in tableaux:
#         if not len(payeur):
#             payeur = 'Département'
#         if payeur not in cahier.sheetnames:
#             nouvelle_feuille = cahier.copy_worksheet(cahier['Modèle'])
#             nouvelle_feuille.title = payeur

#         heures = dataframe_to_rows(
#             heures.loc[:, ['Date',
#                            'Description des travaux effectués',
#                            "Nbr d'heures"]],
#             index=False,
#             header=False)
#         feuille = cahier[payeur]
#         for date, desc, nbr in itertools.chain(
#                 feuille.iter_rows(min_row=19,
#                                   max_row=32,
#                                   min_col=1,
#                                   max_col=3),
#                 feuille.iter_rows(min_row=19,
#                                   max_row=32,
#                                   min_col=4,
#                                   max_col=6)):
#             if not date.value:
#                 try:
#                     date.value, desc.value, nbr.value = next(heures)
#                 except StopIteration:
#                     break

#         # Vérifier si la feuille est pleine, auquel cas il en faut une
#         # nouvelle. On ne devrait pas avoir besoin de faire ça plus qu'une
#         # fois (par mois).
#         heures_restantes: Union[Iterator[Any], list[Any]] =\
#             [h for h in heures]
#         if heures_restantes:
#             nouvelle_feuille = cahier.copy_worksheet(cahier['Modèle'])
#             nouvelle_feuille.title = f'{payeur} 2'

#             heures_restantes = iter(heures_restantes)
#             for date, desc, nbr in itertools.chain(
#                     nouvelle_feuille.iter_rows(min_row=19,
#                                                max_row=32,
#                                                min_col=1,
#                                                max_col=3),
#                     nouvelle_feuille.iter_rows(min_row=19,
#                                                max_row=32,
#                                                min_col=4,
#                                                max_col=6)):

#                 try:
#                     date.value, desc.value, nbr.value = next(heures_restantes)
#                 except StopIteration:
#                     break

#     cahier.save(fichier)

#     return tableaux


def màj_calendrier_temps(présences: pd.DataFrame,
                         fichier: Path = racine / CalendrierTemps):
    pass


if __name__ == '__main__':
    nouveaux = Path.home() / 'Library' / 'Mobile Documents' / \
        'iCloud~is~workflow~my~workflows' / 'Documents'
    boite_de_dépôt = nouveaux / 'Journal de bord/'
    archive = boite_de_dépôt / 'Heures comptées'
    fichiers_textes = list(boite_de_dépôt.glob('*.txt'))
    fichiers_photos = list(boite_de_dépôt.glob('*.png')) + \
        list(boite_de_dépôt.glob('*.jpeg'))
    fichiers_tâches_complétées = [
        f for f in fichiers_textes if 'complétée' in str(f)]

    nouvelles_données = extraire(fichiers_tâches_complétées)
    données_formatées = formater(nouvelles_données)
    données_prêtes = màj_temps_technicien(données_formatées)

    # temps_atelier = heures_atelier(données_formatées)
    # données_atelier = màj_temps_atelier(temps_atelier)

    proportions = répartition(données_formatées)
    présences = compte_des_heures(données_formatées)
    données_présences = màj_calendrier_temps(présences)

    moment = Dt.now()
    fichier = racine / f'màj {moment:%Y-%m-%d %H_%M}.xlsx'
    with pd.ExcelWriter(fichier) as excel:
        données_formatées.to_excel(excel, sheet_name='Données')
        proportions.to_excel(excel, sheet_name='Proportions')
        présences.to_excel(excel, sheet_name='Présences')

    archiver(fichiers_textes, fichiers_photos, archive=archive)
